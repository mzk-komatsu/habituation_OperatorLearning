import numpy as np
from functions import *
import matplotlib.pyplot as plt
import sys
from openpyxl import Workbook, load_workbook
import pickle
sys.path.append("../analysis")
import tools
from distutils.util import strtobool
import json

# ------------------------------------------------------
# Execution(N: id of train settings, M: id of prediction settings):
# python main_pred.py pred_settingsM.yaml train_settingsN.yaml
#
# [pred_settingsM.yaml]
#   - sys_name: name of habituating system (e.g., "TysonAL2D")
#   - recov: whether input data set contain repetitive stimuli with recovery stimulus
#              (0: no recovery stimuli, 1: all input functions contain recovery stimuli, 2: some of input functions contain recovery stimuli)
#   - ranges_id: id of ranges to be evaluated (refer to "sys_name".json in data_generation folder) (e.g., [1,2,3])
# ------------------------------------------------------


# Function for saving results
# Save evaluation metrics
def save_eval_result(exids, parent, RESULT, dmodel, sys_name):
    workbook = Workbook()
    for result in (RESULT):
        sheet_name = "pred"+str(exids[0])
        if i == 0:
            sheet = workbook.active
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        for row_idx, row in enumerate(result, start=1):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
    workbook.save(parent+"/result_"+dmodel+".xlsx")
    return

# Function for loading results
# Load evaluation metrics
def load_eval_result(exid, parent, dmodel, sys_name):
    excel_name = parent+"/result_"+dmodel+".xlsx"
    print("\n Load:", excel_name)
    workbook = load_workbook(excel_name)
    sheet = workbook["pred"+str(exid)]
    for row in sheet.iter_rows(values_only=True):
        print(row)
    return

# Specify json and yaml files
with open("../data_generation/path.json", 'r', encoding='utf-8') as f:
    JSON = json.load(f)
global_path ="../data_generation/"+JSON["global_path"]
with open(global_path, 'r', encoding='utf-8') as f:
    gJSON = json.load(f)
ht_thres = gJSON["ht_thres"]

args = sys.argv
print(args)
if len(args) == 1:
    eval_path ="pred_settings0.yaml"
    with open(eval_path, 'rb') as f:
        eyml = yaml.safe_load(f)
    train_path = "train_settings0.yaml"
else:
    eval_path = args[1]
    with open(eval_path, 'rb') as f:
        eyml = yaml.safe_load(f)
    train_path = args[2]
    if len(args) > 3:
        hpo = bool(strtobool(args[3]))
    else:
        hpo = False

sys_name = eyml["sys_name"]
p_recov = eyml["recov"]
ranges_id = eyml["ranges_id"]
train_exid = train_path[-6] # id of training settings
pred_exid = eval_path[-6]  # id of prediction settings
exids = [pred_exid, train_exid]

sys_json_path = "../data_generation/"+sys_name+".json"
with open(sys_json_path, 'r', encoding='utf-8') as f:
  sys_JSON = json.load(f)
ranges = sys_JSON["ranges"]
num_dp = sys_JSON["num_dp"]

# Load and specify settings
ep, sub_int, model_name, _, model_id, modes, recov, _, _, _, L, width = get_learn_settings(train_path)
F, T, num_dp, ntrain, ntest, bs, grids, f_mean, f_var, numi, numo, pred_ntrain, pred_ntest, split = get_settings(train_path, sys_json_path, global_path)
T = T*split
num_data = pred_ntrain +pred_ntest
ntrain = pred_ntrain
ntest = pred_ntest
num_display = 1
if model_name == "donet":
    from functions_donet import * 
    
# Update hyperparameters if hpo == True
if hpo == True:
    d = np.load("HPO/train"+str(exids[1])+"_"+model_name+".npy")
# Prediction varying ranges of frequency of repetitive stimuli
RESULT = []
res = np.zeros((sub_int+1,2*len(ranges_id)))
for pid, p in enumerate(ranges_id):
  fm = ranges[p][0]
  fv = ranges[p][1]
  freq_name = str(fm)+"-"+str(fv)
  for i in range(sub_int+1):
      if i == 0: # Choose whether display the name of datasets or not
          print("\n")
          display_data = True
      else:
          display_data = False
      if model_name == "donet":
          f_var = 0.5
          sub = 2**i
          if hpo == True:
            L = d[0]
            width = d[1]
          x_train, y_train, x_test, y_test= load_data_donet(sys_name, ntrain=ntrain, ntest=ntest, num_dp = num_dp, sub=sub, batch_size=bs, freq_mean = f_mean, freq_var = f_var, Flevel =F, numi=numi, numo=numo, recov=recov, display_data = True)
          mse_test, l2_test, num_dmodel_params, y_pred = predict_donet(exids, sys_name, x_train, y_train, x_test, y_test, T, ep, num_dp, sub, sub_int, L, width, recov, hpo = False)
      else:
          if hpo == True:
            L = d[0]
            width = d[1]
            modes = d[2]
          train_loader, test_loader, x_train, y_train, x_test, y_test,params, Fs = load_data(sys_name, ntrain=ntrain, ntest=ntest, num_dp = num_dp, sub=2**i,freq_mean = fm, freq_var = fv, Flevel =F, T =T, rand=False,numi=numi, numo=numo, recov=p_recov, batch_size = bs, display_data = display_data)
          if model_name == "fno":
            pred, pred_ht, mse_test, l2_test, num_dmodel_params = predict_plot(exids, sys_name, num_data, x_test, y_test, test_loader, bs, params, Fs, 2**sub_int, i, freq_name, F,  T, model_id, model_name,  num_dp = num_dp, epochs = ep, L = L, modes = modes, width = width, num_display = num_display, numi=numi, numo=numo, hpo = hpo)
          elif model_name == "lno":
            pred, pred_ht, mse_test, l2_test, num_dmodel_params = predict_plot(exids, sys_name, num_data, x_test, y_test, test_loader, bs, params, Fs, 2**sub_int, i, freq_name, F,  T, model_id, model_name, num_dp = num_dp, epochs = ep, L = L, modes = modes, width = width, num_display = num_display, numi=numi, numo=numo, hpo = hpo)
      res[i, pid*2] = mse_test
      res[i, pid*2+1] = l2_test

RESULT.append(res)
parent = "train"+str(exids[1])+"_"+"pred"+str(exids[0])+"_"+(model_name)+"_"+sys_name
save_eval_result(exids, parent, RESULT, model_name, sys_name)
load_eval_result(exids[0], parent, model_name, sys_name)
